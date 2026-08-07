#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Автообновление данных калькулятора «Дмей автала».

ИСТОЧНИКИ (цепочка, робот идёт пока не получит валидный ответ):
  A. Битуах Леуми — страница размеров пособия (через текстовый прокси,
     сайт отдаёт SharePoint-разметку и закрыт для прямого парсинга)
  B. Коль Зхут — резерв, там та же официальная таблица ставок и потолков
  C. Последние сохранённые значения в JSON

Дополнительно:
  - Минимальная зарплата: официальный XLSX Битуах Леуми с data.gov.il

Ни одно значение не перезаписывается «вслепую»: есть проверки диапазонов,
при подозрительных числах сохраняется прежнее и пишется предупреждение.
"""

import datetime
import json
import pathlib
import re
import sys
import urllib.request

BASE = pathlib.Path(__file__).resolve().parent.parent
DATA = BASE / "data" / "avtala.json"
REPORT = BASE / "data" / "last-run.json"

UA = {"User-Agent": "avtala-databot/1.0 (+https://exodus611.github.io/finance-portal/avtala.html; открытые данные, 1 запрос в сутки)"}
PROXY = "https://r.jina.ai/"

BTL_RATES = ("https://www.btl.gov.il/benefits/Unemployment/Pages/"
             "%D7%A9%D7%99%D7%A2%D7%95%D7%A8%D7%99%20%D7%93%D7%9E%D7%99%20%D7%90%D7%91%D7%98%D7%9C%D7%94.aspx")
KOLZCHUT = "https://www.kolzchut.org.il/he/%D7%93%D7%9E%D7%99_%D7%90%D7%91%D7%98%D7%9C%D7%94"
MINWAGE_XLSX = "https://www.btl.gov.il/Mediniyut/GeneralData/Documents/sharminimum.xlsx?Web=1"

notes, warnings, sources_used = [], [], []


def fetch(url, timeout=70, binary=False):
    req = urllib.request.Request(url, headers=UA)
    data = urllib.request.urlopen(req, timeout=timeout).read()
    return data if binary else data.decode("utf-8", "replace")


# ───────────── потолки и ставки ─────────────

def parse_caps_and_brackets(text):
    """Достаёт дневные потолки и таблицу ступеней.

    Числа ищутся только рядом с ключевыми словами — иначе в потолки попадают
    суммы из поясняющих примеров на странице.
    """
    out = {}

    # Потолок: ищем только во фрагменте про максимальный размер
    m = re.search(r"(?:הסכום\s+המירבי|הסכום\s+המרבי|תקרה)(.{0,700})", text, re.S)
    zone = m.group(1) if m else ""
    # «в первые 125 дней ... X ₪» и «в остальные дни ... Y ₪»
    first = re.search(r"125\s*הימים\s*הראשונים.{0,160}?(\d{3}\.\d{2})\s*₪", zone, re.S)
    rest = re.search(r"(?:בשאר\s+ימי\s+הזכאות|מהיום\s+ה-?126).{0,160}?(\d{3}\.\d{2})\s*₪", zone, re.S)
    if first and rest:
        a, b = float(first.group(1)), float(rest.group(1))
        if 200 <= b < a <= 900:
            out["daily_first_period"] = a
            out["daily_after_125"] = b

    # Ступени: только внутри таблицы, где рядом стоят обе ставки
    rows = re.findall(
        r"(\d[\d,]*\.?\d*)\s*₪\s*\|[^|]{0,140}?(\d{2})%\s*\|[^|]{0,140}?(\d{2})%", text)
    brackets = []
    for edge, r_under, r_over in rows:
        edge_v = float(edge.replace(",", ""))
        if not (100 <= edge_v <= 5000):
            continue
        ru, ro = int(r_under) / 100.0, int(r_over) / 100.0
        if ru > ro:                      # до 28 лет ставка не может быть выше
            continue
        brackets.append({"upto": edge_v, "rate_under28": ru, "rate_28plus": ro})

    seen, uniq = set(), []
    for b in sorted(brackets, key=lambda x: x["upto"]):
        if b["upto"] in seen:
            continue
        seen.add(b["upto"])
        uniq.append(b)
    if len(uniq) >= 3:
        out["brackets"] = uniq
    return out


def source_btl():
    txt = fetch(PROXY + BTL_RATES, timeout=90)
    if len(txt) < 1500:
        raise ValueError("страница слишком короткая (%d байт)" % len(txt))
    res = parse_caps_and_brackets(txt)
    if not res:
        raise ValueError("не удалось разобрать ставки")
    return res


def source_kolzchut():
    txt = fetch(PROXY + KOLZCHUT, timeout=90)
    if len(txt) < 5000:
        raise ValueError("страница слишком короткая")
    res = parse_caps_and_brackets(txt)
    if not res:
        raise ValueError("не удалось разобрать ставки")
    return res


# ───────────── минимальная зарплата ─────────────

def fetch_min_wage():
    raw = fetch(MINWAGE_XLSX, timeout=90, binary=True)
    tmp = pathlib.Path("/tmp/_minwage.xlsx")
    tmp.write_bytes(raw)
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("нет openpyxl")
    wb = openpyxl.load_workbook(tmp, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Первая строка данных — самая свежая дата. Колонки: дата, ..., месячная (индекс 5), часовая (4)
    for row in ws.iter_rows(min_row=4, max_row=8, values_only=True):
        if not row or not row[0]:
            continue
        date_s = str(row[0]).strip()
        m = re.match(r"(\d{1,2})\.(\d{2})\.(\d{4})", date_s)
        if not m:
            continue
        since = "%s-%s-%02d" % (m.group(3), m.group(2), int(m.group(1)))
        monthly = float(row[5])
        hourly = float(row[4])
        return {"monthly": monthly, "hourly": hourly, "since": since}
    raise ValueError("не нашёл строку с данными")


# ───────────── сборка ─────────────

def main():
    d = json.loads(DATA.read_text(encoding="utf-8"))
    today = datetime.date.today().isoformat()

    parsed = None
    for name, fn in (("Битуах Леуми", source_btl), ("Коль Зхут (резерв)", source_kolzchut)):
        try:
            parsed = fn()
            sources_used.append(name)
            notes.append("%s: получены ставки и потолки." % name)
            break
        except Exception as e:
            warnings.append("%s недоступен (%s)." % (name, e))

    if parsed:
        caps = d["caps"]
        nf = parsed.get("daily_first_period")
        na = parsed.get("daily_after_125")
        if nf and na and 200 <= na < nf <= 900 and (nf - na) > 50:
            if abs(nf - caps["daily_first_period"]) > 0.001 or abs(na - caps["daily_after_125"]) > 0.001:
                notes.append("Потолки обновлены: %.2f → %.2f и %.2f → %.2f ₪/день."
                             % (caps["daily_first_period"], nf, caps["daily_after_125"], na))
            caps["daily_first_period"] = nf
            caps["daily_after_125"] = na
        elif nf or na:
            warnings.append("Потолки выглядят подозрительно (%s / %s) — сохранены прежние." % (nf, na))

        nb = parsed.get("brackets")
        if nb and len(nb) == len(d["brackets"]):
            valid = all(0.1 <= b["rate_under28"] <= 0.95 and 0.1 <= b["rate_28plus"] <= 0.95
                        and b["rate_under28"] <= b["rate_28plus"] for b in nb)
            if valid:
                if nb != d["brackets"]:
                    notes.append("Таблица ступеней обновлена.")
                d["brackets"] = nb
            else:
                warnings.append("Ступени не прошли проверку — сохранены прежние.")
        elif nb:
            warnings.append("Число ступеней изменилось (%d вместо %d) — требуется сверка вручную."
                            % (len(nb), len(d["brackets"])))
    else:
        warnings.append("Все источники ставок недоступны — используются последние сохранённые значения.")

    try:
        mw = fetch_min_wage()
        if 4000 <= mw["monthly"] <= 12000:
            if abs(mw["monthly"] - d["macro"].get("minimum_wage_monthly", 0)) > 0.01:
                notes.append("Минимальная зарплата: %.2f → %.2f ₪ (с %s)."
                             % (d["macro"].get("minimum_wage_monthly", 0), mw["monthly"], mw["since"]))
            d["macro"]["minimum_wage_monthly"] = mw["monthly"]
            d["macro"]["minimum_wage_hourly"] = mw["hourly"]
            d["macro"]["minimum_wage_since"] = mw["since"]
            sources_used.append("data.gov.il — минимальная зарплата")
        else:
            warnings.append("Минимальная зарплата вне разумного диапазона (%s)." % mw["monthly"])
    except Exception as e:
        warnings.append("Минимальная зарплата недоступна (%s)." % e)

    d["updated"] = today
    d["updated_by"] = "auto"
    d["auto"] = {
        "last_run": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
        "ok": not warnings,
        "sources_used": sources_used,
        "notes": notes,
        "warnings": warnings,
    }

    DATA.write_text(json.dumps(d, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    REPORT.write_text(json.dumps(d["auto"], ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print("=== ИСТОЧНИКИ ===")
    for s in sources_used:
        print("  ✓", s)
    print("=== ЗАМЕТКИ ===")
    for n in notes:
        print("  •", n)
    if warnings:
        print("=== ПРЕДУПРЕЖДЕНИЯ ===")
        for w in warnings:
            print("  !", w)

    return 1 if not sources_used else 0


if __name__ == "__main__":
    sys.exit(main())